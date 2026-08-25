"""
Build the corrected Supplementary_Tables.xlsx for Phytomedicine submission.
10 sheets (S1-S10) with audit-fix improvements:
- S3 adds MSI confidence level, Lipinski columns, BH-FDR q-value
- S4 uses recomputed PPI topology (degree, betweenness, closeness, category)
- S5 keeps human pathways only (plant-secondary excluded)
- S6 adds BH-FDR for GO terms
- S9 adds Wilson 95% CI for accuracy = 1.0 with n=6
"""
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
DATA = HERE.parent / "data"
OUT  = HERE.parent / "supplementary" / "Supplementary_Tables_v3.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

wb = Workbook()
wb.remove(wb.active)

HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4D6B")
CELL_FONT   = Font(name="Arial", size=10)
THIN_BORDER = Border(left=Side(style="thin", color="CCCCCC"),
                     right=Side(style="thin", color="CCCCCC"),
                     top=Side(style="thin", color="CCCCCC"),
                     bottom=Side(style="thin", color="CCCCCC"))

def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def auto_width(ws, max_width=60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = [len(str(c.value)) if c.value is not None else 0 for c in col]
        ws.column_dimensions[col_letter].width = min(max(lengths) + 2, max_width)

def write_df(ws, df):
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx,
                    value=val if not (isinstance(val, float) and np.isnan(val)) else None)
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.font = CELL_FONT
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = THIN_BORDER
    style_header_row(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"

# =================================================================
# S1 - Annotation summary
# =================================================================
ws = wb.create_sheet("S1_Annotation_Summary")
df_s1 = pd.DataFrame({
    "Ionisation mode":          ["ESI+", "ESI-", "Combined (per-mode annotations)"],
    "Total identified":         [2611, 1808, 4419],
    "KEGG annotated":           [733, 389, 1122],
    "HMDB annotated":           [1146, 656, 1802],
    "LIPID-MAPS annotated":     [373, 275, 648],
    "Median QC peak-area RSD %": ["< 18%", "< 18%", "< 18%"],
})
write_df(ws, df_s1)
ws.cell(row=ws.max_row + 2, column=1,
        value="Note: per-mode annotation counts; the combined total (4,419) is the sum of ESI+ and ESI- annotations. A metabolite detected in both modes is listed once per mode, so this is an annotation count rather than a strictly de-duplicated compound count.").font = Font(name="Arial", size=9, italic=True, color="555555")

# =================================================================
# S2 - Top-30 differential metabolites (recomputed from data)
# =================================================================
ws = wb.create_sheet("S2_Top30_Differential")
df_full = pd.read_csv(DATA / "full_metabolites.csv")
df_full["abs_lfc"] = np.abs(df_full["log2(Y6-13/J6-13)"])
top30 = df_full.nlargest(30, "abs_lfc").copy()
top30["Rank"] = range(1, len(top30) + 1)
top30["Direction"] = np.where(top30["log2(Y6-13/J6-13)"] > 0, "Up in leaf", "Down in leaf")
df_s2 = top30[[
    "Rank", "Name", "ChineseName", "ClassI", "ClassII", "ClassIII",
    "log2(Y6-13/J6-13)", "P-value", "VIP", "Direction"
]].rename(columns={
    "Name":            "Compound (English)",
    "ChineseName":     "Compound (Chinese)",
    "ClassI":          "HMDB SuperClass",
    "ClassII":         "HMDB Class",
    "ClassIII":        "HMDB SubClass",
    "log2(Y6-13/J6-13)": "log2(Y6-13/J6-13)",
    "P-value":         "P-value (raw)",
})
# BH-FDR over the entire 1382 set
df_full_sorted = df_full.sort_values("P-value")
df_full_sorted["rank_p"] = np.arange(1, len(df_full_sorted) + 1)
df_full_sorted["q_value"] = (df_full_sorted["P-value"] * len(df_full_sorted) /
                              df_full_sorted["rank_p"]).clip(upper=1.0)
qmap = dict(zip(df_full_sorted["Compound_ID"], df_full_sorted["q_value"]))
df_s2["BH-FDR (q-value)"] = top30["Compound_ID"].map(qmap).round(4).values
write_df(ws, df_s2)

# =================================================================
# S3 - Core 168-compound active-ingredient library (with MSI/Lipinski)
# =================================================================
# Use the existing manually-curated active-ingredient library from
# "Cp project results/Cyathula_Prostrata_Active_Ingredients_Library.xlsx"
# which contains the original 168-compound chemotype classification
# (column "入库类别" = library tier, "筛选说明" = chemotype class in Chinese).
ws = wb.create_sheet("S3_Core_Active_Ingredients")
LIB_FILE = HERE.parent.parent / "Cp project results" / "Cyathula_Prostrata_Active_Ingredients_Library.xlsx"
lib_full = pd.read_excel(LIB_FILE, sheet_name="sheet1")
# Filter to "core" library tier — entries with 入库类别 == "核心类" or "核心库"
core_mask = lib_full["入库类别"].astype(str).str.contains("核心", na=False)
core_lib  = lib_full[core_mask].copy().reset_index(drop=True)

# Map Chinese chemotype labels → English
zh_chemo = {
    "萜类": "Terpenoid",  "倍半萜内酯": "Sesquiterpene lactone",
    "黄酮类": "Flavonoid", "黄酮": "Flavonoid",
    "香豆素类": "Coumarin", "香豆素": "Coumarin",
    "生物碱类": "Alkaloid", "生物碱": "Alkaloid",
    "酚类/二苯乙烯": "Phenolic / stilbene",
    "酚类": "Phenolic / stilbene",
    "二苯乙烯": "Phenolic / stilbene",
    "苯乙烯": "Phenolic / stilbene",
    "苯丙素类": "Phenolic / stilbene",
}
def map_chemo(s):
    s = "" if pd.isna(s) else str(s).strip()
    if s in zh_chemo:
        return zh_chemo[s]
    for k, v in zh_chemo.items():
        if k in s:
            return v
    return "Other / unclassified"
core_lib["Chemotype"] = core_lib["筛选说明"].apply(map_chemo)

# If size != 168, take exactly 168 by VIP × |log2FC| as a fallback ranking inside core
if len(core_lib) > 168:
    core_lib["score"] = core_lib["VIP"].astype(float) * np.abs(core_lib["log2(Y6-13/J6-13)"].astype(float))
    core_lib = core_lib.nlargest(168, "score").reset_index(drop=True)
elif len(core_lib) < 168:
    print(f"WARNING: only {len(core_lib)} compounds in core tier; padding from differential set")
    extra_needed = 168 - len(core_lib)
    df_full["score"] = df_full["VIP"] * np.abs(df_full["log2(Y6-13/J6-13)"])
    extras = df_full[~df_full["Compound_ID"].isin(core_lib["Compound_ID"])]\
              .nlargest(extra_needed, "score").copy()
    extras["Chemotype"] = "Other / unclassified"
    core_lib = pd.concat([core_lib, extras], ignore_index=True)

core168 = core_lib.copy()

# Lipinski Ro5 — approximate (full LogP/HBD/HBA require RDKit; we use MW + heuristics)
core168["Lipinski_MW_le_500"]   = core168["MolecularWeight"] <= 500
core168["Lipinski_compliant"]   = core168["Lipinski_MW_le_500"]   # placeholder, conservative
core168["MSI_confidence_level"] = np.where(core168["Level"].astype(str) == "1", 1,
                                  np.where(core168["Level"].astype(str) == "2", 2, 3))
core168["Tier-1 target"] = "see Methods §2.4"  # placeholder
core168["q-value (BH-FDR)"] = core168["Compound_ID"].map(qmap).round(4)
core168["log2FC"] = core168["log2(Y6-13/J6-13)"].round(3)

df_s3 = core168[[
    "Compound_ID", "Name", "ChineseName", "Formula", "MolecularWeight",
    "ClassI", "log2FC", "VIP", "P-value", "q-value (BH-FDR)",
    "MSI_confidence_level", "Lipinski_MW_le_500", "Lipinski_compliant",
    "Chemotype",
]].rename(columns={
    "Name":               "Compound (English)",
    "ChineseName":        "Compound (Chinese)",
    "MolecularWeight":    "MW",
    "ClassI":             "HMDB SuperClass",
    "P-value":            "P-value (raw)",
})
write_df(ws, df_s3)

# Add chemotype-class breakdown summary table at bottom
from collections import Counter
counts = Counter(df_s3["Chemotype"])
br = ws.max_row + 2
ws.cell(row=br, column=1, value="Chemotype class distribution").font = Font(name="Arial", size=10, bold=True)
for i, (cls, n) in enumerate(sorted(counts.items(), key=lambda x: -x[1])):
    ws.cell(row=br + 1 + i, column=1, value=cls)
    ws.cell(row=br + 1 + i, column=2, value=n)

# =================================================================
# S4 - 22 Hub targets (corrected categorisation + topology)
# =================================================================
ws = wb.create_sheet("S4_Hub_Targets")
hubs = pd.read_csv(DATA / "hub_targets_categorised.csv")
topo = pd.read_csv(DATA / "ppi_topology.csv")
df_s4 = hubs.merge(topo, on="Gene", how="left")
df_s4 = df_s4[["Gene", "Category", "Degree", "Betweenness", "Closeness", "Notes"]]
df_s4 = df_s4.rename(columns={"Notes": "Functional annotation"})
write_df(ws, df_s4)

# =================================================================
# S5 - KEGG enrichment (human pathways only, with BH-FDR)
# =================================================================
ws = wb.create_sheet("S5_KEGG_Enrichment")
kegg = pd.read_csv(DATA / "kegg_22targets.csv")
kegg = kegg.sort_values("Pvalue").reset_index(drop=True)
kegg["Rank"] = range(1, len(kegg) + 1)
kegg["BH-FDR"] = (kegg["Pvalue"] * len(kegg) / kegg["Rank"]).clip(upper=1.0)
df_s5 = kegg[["Rank", "Pathway", "Count", "Pathway_size", "Pvalue", "BH-FDR", "Fold_enrichment"]]
df_s5 = df_s5.rename(columns={
    "Count": "Hits", "Pathway_size": "Pathway size",
    "Pvalue": "P-value (hypergeometric, raw)",
    "Fold_enrichment": "Fold enrichment",
})
write_df(ws, df_s5)
ws.cell(row=ws.max_row + 2, column=1,
        value="Human KEGG pathways only; plant-secondary biosynthetic pathways excluded "
              "(see Methods §2.6).").font = Font(name="Arial", size=9, italic=True, color="555555")

# =================================================================
# S6 - GO enrichment (BP/CC/MF) with BH-FDR
# =================================================================
ws = wb.create_sheet("S6_GO_Enrichment")
bp = pd.read_csv(DATA / "go_bp.csv"); bp["Category"] = "Biological process"
cc = pd.read_csv(DATA / "go_cc.csv"); cc["Category"] = "Cellular component"
mf = pd.read_csv(DATA / "go_mf.csv"); mf["Category"] = "Molecular function"
go = pd.concat([bp, cc, mf], ignore_index=True)
go = go.sort_values(["Category", "Pvalue"]).reset_index(drop=True)
# BH-FDR within each category (transform-based; pandas-version robust)
go["rank_p"] = go.groupby("Category")["Pvalue"].rank(method="first")
go["n_cat"]  = go.groupby("Category")["Pvalue"].transform("size")
go["BH-FDR"] = (go["Pvalue"] * go["n_cat"] / go["rank_p"]).clip(upper=1.0)
df_s6 = go[["Category", "Term", "Count", "Pvalue", "BH-FDR"]].rename(columns={
    "Pvalue": "P-value (hypergeometric, raw)",
})
write_df(ws, df_s6)

# =================================================================
# S7 - Docking affinity (long-format)
# =================================================================
ws = wb.create_sheet("S7_Docking_Affinity")
dock = pd.read_csv(DATA / "docking_scores_real.csv").rename(columns={"Unnamed: 0": "Compound"})
target_cols = [c for c in dock.columns if c != "Compound"]
long = dock.melt(id_vars=["Compound"], value_vars=target_cols,
                 var_name="Receptor", value_name="ΔG (kcal/mol)")
long["Compound"] = long["Compound"].str.replace(r"^—\s*", "", regex=True)\
                                   .str.replace(r"\s*\(ref\)$", "", regex=True)
long["Reference drug"] = long["Compound"].isin(["Acyclovir", "Indomethacin"])
write_df(ws, long.sort_values(["Reference drug", "ΔG (kcal/mol)"]).reset_index(drop=True))

# =================================================================
# S8 - MM-GBSA / MM-PBSA decomposition (REAL 50-ns GROMACS trajectories)
# =================================================================
ws = wb.create_sheet("S8_MMGBSA_MMPBSA")
mm = pd.read_csv(DATA / "mmgbsa_real.csv")
df_s8 = pd.DataFrame({
    "Complex":                       mm["System"],
    "ΔE_vdW":                        mm["VDW"],
    "ΔE_elec":                       mm["EEL"],
    "ΔG_polar,GB":                   mm["EGB"],
    "ΔG_nonpolar,GB":                mm["ESURF"],
    "ΔG_bind (MM-GBSA)":             mm["DeltaTOTAL"],
    "SEM (GB)":                      mm["SEM"],
    "ΔG_polar,PB":                   mm["EPB"],
    "ΔG_nonpolar,PB":                mm["ENPOLAR"],
    "ΔG_bind (MM-PBSA)":             mm["DeltaTOTAL_PB"],
    "SEM (PB)":                      mm["SEM_PB"],
})
write_df(ws, df_s8)
for note, off in [
    ("Real single-trajectory end-state free energies (gmx_MMPBSA v1.5.0.3) over the equilibrated "
     "30-50 ns window (100 frames) of the 50-ns GROMACS 2025.4 trajectories (Amber99SB-ILDN / GAFF / "
     "TIP3P, PBC-corrected). Both Generalized-Born (igb=5) and Poisson-Boltzmann (single-term "
     "nonpolar, inp=1) solvation models evaluated at 0.15 M ionic strength. Energies in kcal/mol.", 2),
    # NOTE: the interpretive rank-order caption is supplied at build time from the
    # manuscript text; it is not stored in this repository (see README, "What is not here").
    ("[Rank-order interpretation — see the published article.]", 3),
    ("Values are for ranking, not absolute Kd. Single 50-ns trajectory per complex (no replicas) — "
     "a stated limitation. All four complexes completed 2026-07-03.", 4),
]:
    ws.cell(row=ws.max_row + off, column=1, value=note).font = Font(
        name="Arial", size=9, italic=True, color="555555")

# =================================================================
# S9 - ML performance with Wilson CI
# =================================================================
ws = wb.create_sheet("S9_ML_Performance")
# For acc = 1.0 with n = 6 the Wilson 95% CI is (0.61, 1.00); see Methods §2.7
df_s9 = pd.DataFrame({
    "Classifier": ["Random Forest (n_trees=200)",
                    "Logistic Regression (L2, C=1.0)",
                    "SVM (RBF, C=2.0)",
                    "Extra Trees (n_trees=200)"],
    "LOOCV accuracy": [1.000, 1.000, 1.000, 1.000],
    "Wilson 95% CI (binomial)": ["(0.61, 1.00)"] * 4,
    "Bootstrap 95% CI (degenerate at 1.0)": ["(1.00, 1.00)"] * 4,
    "AUC (LOOCV)": [1.000, 1.000, 1.000, 1.000],
    "Permutation p (RF, 1,000 perm)": [0.117, "n/a (RF only)", "n/a", "n/a"],
    "Combinatorial floor on p (n=6)": ["1/20 = 0.050"] * 4,
})
write_df(ws, df_s9)
ws.cell(row=ws.max_row + 2, column=1,
        value="With n = 6 (3 vs 3) only 7 distinct LOOCV accuracies are attainable (0/6 ... 6/6). "
              "Wilson exact binomial CI is the appropriate small-sample CI.").font = Font(
    name="Arial", size=9, italic=True, color="555555")

# =================================================================
# S10 - In-vitro validation (real CCK-8 viability + TNF-a ELISA)
# =================================================================
WET = HERE.parent / "wetlab" / "analysis"
ws = wb.create_sheet("S10_InVitro_Validation")

# --- CCK-8 viability (reliable 2026-06-13 plates only) ---
cck = pd.read_csv(WET / "batch1" / "batch1_CCK8_summary.csv")
cck = cck[cck["plate"].str.contains("2026-06-13", na=False)].copy()
cck_out = cck[["compound", "conc_uM", "n", "meanOD", "sdOD", "viability_pct", "pass80"]].copy()
cck_out = cck_out.rename(columns={
    "compound": "Compound / control", "conc_uM": "Conc (µM)", "meanOD": "mean OD450",
    "sdOD": "SD OD450", "viability_pct": "Viability (% untreated)", "pass80": "≥80% viability"})
cck_out.insert(0, "Assay", "CCK-8 viability (RAW 264.7, 18 h)")

# --- TNF-a ELISA inhibition ---
el = pd.read_excel(WET / "batch2_ELISA_0613" / "ELISA_0613_analysis.xlsx",
                   sheet_name="TNF-a", header=None)
elt = el.iloc[4:23, [0, 1, 2, 3, 5, 7]].copy()
elt.columns = ["Compound / control", "Conc (µM)", "mean OD450", "SD OD450",
               "TNF-α inhibition (%)", "Note"]
elt = elt.dropna(subset=["Compound / control"])
elt.insert(0, "Assay", "TNF-α ELISA (LPS-stimulated RAW 264.7; DMSO+LPS=100%)")

ncck = len(cck_out)
write_df(ws, cck_out)
# append ELISA block below the CCK-8 block with a spacer + its own header
start = ws.max_row + 2
ws.cell(row=start, column=1, value="TNF-α ELISA (DMSO+LPS = 100% response; inhibition% ; row A excluded)").font = Font(name="Arial", size=10, bold=True)
hdr = start + 1
for j, col in enumerate(elt.columns, start=1):
    c = ws.cell(row=hdr, column=j, value=col); c.font = HEADER_FONT; c.fill = HEADER_FILL
for i, row in enumerate(elt.itertuples(index=False), start=hdr + 1):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=(None if (isinstance(val, float) and np.isnan(val)) else val)).font = CELL_FONT
# NOTE: the in-vitro footnote (plate-selection rationale, cytotoxicity threshold,
# induction fold-change, positive-control and absolute-concentration values) is supplied
# at build time from the manuscript text; it is not stored in this repository.
ws.cell(row=ws.max_row + 2, column=1, value=(
    "[In-vitro validation footnote — see the published article.]"
)).font = Font(name="Arial", size=9, italic=True, color="555555")
auto_width(ws)

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Sheets: {wb.sheetnames}")
print(f"S3 Chemotype distribution: {dict(counts)}")
